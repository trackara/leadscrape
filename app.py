"""
app.py — Lead Finder · San Canzian
Run:  streamlit run app.py
"""

import io
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pipeline import run_pipeline, enrich_urls, leads_to_records
from config import (
    SEGMENT_QUERIES, DEFAULT_TARGET_KEYWORDS, ANTI_KEYWORDS,
    EXISTING_PARTNERS, DOMAIN_BLACKLIST, UI_STRINGS, SUGGESTED_COMPETITORS,
)
from modules import dedup_db, competitor_mining, directories, templates

st.set_page_config(page_title="Lead Finder · San Canzian", layout="wide")

# Animated intro overlay — escapes Streamlit iframe, injects into parent page
import streamlit.components.v1 as components
from intro_html import INTRO_HTML
components.html(INTRO_HTML, height=0)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Serif+Display&display=swap');

@keyframes fadeUp   { from { opacity:0; transform:translateY(14px); } to { opacity:1; transform:translateY(0); } }
@keyframes fadeIn   { from { opacity:0; } to { opacity:1; } }
@keyframes slideIn  { from { opacity:0; transform:translateX(-10px); } to { opacity:1; transform:translateX(0); } }
@keyframes countUp  { from { opacity:0; transform:scale(0.85); } to { opacity:1; transform:scale(1); } }
@keyframes shimmer  { 0%,100%{opacity:1} 50%{opacity:0.6} }
@keyframes barGrow  { from { width:0; } to { width:100%; } }

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }

[data-testid="stAppViewContainer"] { background: #F2EFE9; }
[data-testid="stHeader"] { background: transparent !important; }
[data-testid="stSidebar"] {
    background: #1B3A2D !important;
    border-right: none !important;
}
[data-testid="stSidebar"] * { color: #C8D9C2 !important; }
[data-testid="stSidebar"] h3 {
    font-size: 10px !important; font-weight: 600 !important;
    letter-spacing: 0.12em; text-transform: uppercase;
    color: #7A9B84 !important; margin: 0 0 8px;
}
[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] div[role="slider"] { background: #4CAF80 !important; }
[data-testid="stSidebar"] [data-baseweb="input"] { background: rgba(255,255,255,0.08) !important; border-color: rgba(255,255,255,0.1) !important; color: #fff !important; }
[data-testid="stSidebar"] textarea { background: rgba(255,255,255,0.08) !important; border-color: rgba(255,255,255,0.2) !important; color: #FFFFFF !important; font-size: 13px !important; }
[data-testid="stSidebar"] label { color: #a8c4ae !important; }
[data-testid="stSidebar"] [data-baseweb="select"] > div { background: rgba(255,255,255,0.08) !important; border-color: rgba(255,255,255,0.15) !important; }
[data-testid="stSidebar"] [data-baseweb="select"] span { color: #e0ede6 !important; }
[data-testid="stSidebar"] .stCheckbox label span { color: #c8d9c2 !important; }
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; }

#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2.5rem 2.5rem 3rem; max-width: 1200px; }

.sc-label {
    font-size: 10px; font-weight: 600; letter-spacing: 0.14em;
    text-transform: uppercase; color: #9A8F80; margin: 0 0 3px;
    animation: fadeIn 0.5s ease forwards;
}
.sc-title {
    font-family: 'DM Serif Display', serif !important;
    font-size: 2.4rem; font-weight: 400; color: #1A1A18;
    margin: 0 0 1.8rem; line-height: 1.1;
    animation: fadeUp 0.6s ease forwards;
}
.sc-accent-line {
    height: 3px; width: 48px;
    background: linear-gradient(90deg, #2D7A4F, #7BBF8E);
    border-radius: 2px; margin: 6px 0 20px;
    animation: barGrow 0.8s ease forwards;
}

/* Metric cards */
[data-testid="stMetric"] {
    background: #FFFFFF;
    border: none;
    border-radius: 14px;
    padding: 1.1rem 1.3rem 1rem;
    animation: fadeUp 0.5s ease forwards;
    position: relative; overflow: hidden;
}
[data-testid="stMetric"]::before {
    content: ''; position: absolute; top: 0; left: 0;
    width: 100%; height: 3px;
    background: linear-gradient(90deg, #2D7A4F, #7BBF8E);
}
[data-testid="stMetricLabel"] {
    font-size: 10px !important; font-weight: 600 !important;
    letter-spacing: 0.1em; text-transform: uppercase;
    color: #9A8F80 !important;
}
[data-testid="stMetricValue"] {
    font-family: 'DM Serif Display', serif !important;
    font-size: 2.2rem !important; font-weight: 400 !important;
    color: #1A1A18 !important;
    animation: countUp 0.5s ease forwards;
}
[data-testid="stMetricDelta"] { font-size: 11px !important; }

/* nth-child color accents on metric cards */
[data-testid="stHorizontalBlock"] > div:nth-child(1) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#2D7A4F,#7BBF8E); }
[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#059669,#34D399); }
[data-testid="stHorizontalBlock"] > div:nth-child(3) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#D97706,#FCD34D); }
[data-testid="stHorizontalBlock"] > div:nth-child(4) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#2563EB,#93C5FD); }
[data-testid="stHorizontalBlock"] > div:nth-child(5) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#7C3AED,#C4B5FD); }
[data-testid="stHorizontalBlock"] > div:nth-child(6) [data-testid="stMetric"]::before { background: linear-gradient(90deg,#DC2626,#FCA5A5); }

/* Tabs */
[data-testid="stTabs"] [data-baseweb="tab-list"] {
    gap: 0; border-bottom: 1px solid rgba(0,0,0,0.08);
    background: transparent; padding: 0; margin-bottom: 1.5rem;
}
[data-testid="stTabs"] [data-baseweb="tab"] {
    font-size: 13px; font-weight: 500; padding: 10px 22px;
    border-radius: 0; color: #9A8F80 !important;
    border-bottom: 2px solid transparent;
    transition: color 0.2s, border-color 0.2s;
}
[data-testid="stTabs"] [aria-selected="true"] {
    color: #1A1A18 !important; border-bottom: 2px solid #2D7A4F !important;
    background: transparent !important;
}

/* Buttons */
.stButton > button {
    border-radius: 10px !important; font-weight: 500 !important;
    transition: all 0.2s ease !important;
}
.stButton > button[kind="primary"] {
    background: #1B3A2D !important; color: #fff !important;
    border: none !important; font-size: 14px !important;
    letter-spacing: 0.02em !important;
}
.stButton > button[kind="primary"]:hover {
    background: #2D7A4F !important; transform: translateY(-1px) !important;
}
.stButton > button[kind="secondary"]:hover { transform: translateY(-1px) !important; }

[data-testid="stDownloadButton"] > button {
    background: #1B3A2D !important; color: #fff !important;
    border: none !important; border-radius: 10px !important;
    font-weight: 500 !important; transition: all 0.2s !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: #2D7A4F !important; transform: translateY(-1px) !important;
}

/* Inputs */
[data-baseweb="input"], [data-baseweb="textarea"] {
    border-radius: 10px !important;
    transition: border-color 0.2s !important;
}
[data-baseweb="input"]:focus-within, [data-baseweb="textarea"]:focus-within {
    border-color: #2D7A4F !important;
    box-shadow: 0 0 0 3px rgba(45,122,79,0.1) !important;
}

/* Multiselect tags */
[data-baseweb="tag"] { background: #1B3A2D !important; border-radius: 6px !important; }
[data-baseweb="tag"] span { color: #fff !important; }

/* Dataframe */
[data-testid="stDataFrame"] {
    border: 1px solid rgba(0,0,0,0.06) !important;
    border-radius: 14px !important; overflow: hidden;
    animation: fadeUp 0.4s ease forwards;
}

/* Dividers */
hr { border: none; border-top: 1px solid rgba(0,0,0,0.07) !important; margin: 1.5rem 0 !important; }

/* Expanders */
[data-testid="stExpander"] {
    border: 1px solid rgba(0,0,0,0.07) !important;
    border-radius: 12px !important; background: #fff;
}

/* Section description boxes */
.sc-desc {
    background: #fff; border: 1px solid rgba(0,0,0,0.07);
    border-left: 3px solid #2D7A4F;
    border-radius: 0 10px 10px 0;
    padding: 12px 16px; margin: 0 0 1.5rem;
    animation: slideIn 0.4s ease forwards;
}
.sc-desc p { margin:0; font-size:13px; color:#4A4A3F; line-height:1.6; }
.sc-desc strong { color:#1B3A2D; }

/* Radio */
[data-testid="stRadio"] label { font-size: 13px !important; font-weight: 500 !important; }
[data-testid="stRadio"] [data-baseweb="radio"] div[data-checked="true"] { background: #2D7A4F !important; }

/* Form */
[data-testid="stForm"] { border: none !important; padding: 0 !important; }

/* Alerts */
[data-testid="stAlert"] { border-radius: 10px !important; animation: fadeIn 0.3s ease; }
.stSuccess { border-left: 3px solid #059669 !important; }

/* Progress bar */
[data-testid="stProgressBar"] > div > div { background: linear-gradient(90deg, #2D7A4F, #7BBF8E) !important; border-radius: 4px !important; }
</style>
""", unsafe_allow_html=True)

# ---------- Session state ----------
for k, v in {"leads_df": None, "lang": "en", "custom_segments": {}}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ---------- Sidebar ----------
with st.sidebar:
    st.markdown('<p style="font-size:18px;font-weight:600;color:#E8F0E4;margin-bottom:1rem;letter-spacing:-0.01em;">San Canzian</p>', unsafe_allow_html=True)
    lang_choice = st.selectbox("Language / Jezik", ["English", "Hrvatski"])
    st.session_state.lang = "en" if lang_choice == "English" else "hr"
    t = UI_STRINGS[st.session_state.lang]

    st.divider()
    st.subheader("Search depth")
    max_per_query = st.slider("Results per query", 5, 20, 8)
    enrich_opt = st.checkbox("Enrich from websites", value=True)

    st.divider()
    st.subheader("Qualification keywords")
    st.caption("One per line — matches boost score")
    kw_text = st.text_area("kw", value="\n".join(DEFAULT_TARGET_KEYWORDS), height=120, label_visibility="collapsed")
    target_kw = [k.strip() for k in kw_text.splitlines() if k.strip()]

    st.divider()
    st.subheader("Existing partners")
    st.caption("Excluded from results")
    ex_text = st.text_area("ex", value="\n".join(EXISTING_PARTNERS), height=100, label_visibility="collapsed")
    existing = [p.strip() for p in ex_text.splitlines() if p.strip()]

    st.divider()
    st.subheader("Sender")
    sender_name = st.text_input("Name", value="Emma")
    sender_role = st.text_input("Role", value="Sales & Partnerships")

# ---------- Header ----------
st.markdown('<p class="sc-label">San Canzian · Sales dashboard</p>', unsafe_allow_html=True)
st.markdown('<h1 class="sc-title">Lead Finder</h1>', unsafe_allow_html=True)
st.markdown('<div class="sc-accent-line"></div>', unsafe_allow_html=True)

tab_find, tab_outreach, tab_history = st.tabs(["Find Leads", "Outreach", "History"])

# =====================================================================
# TAB 1 — FIND LEADS
# =====================================================================
with tab_find:
    all_segments = {**SEGMENT_QUERIES, **st.session_state.custom_segments}

    mode = st.radio("Source", ["Region search", "Competitor mining", "Directory"], horizontal=True)
    st.write("")

    # --- Region search ---
    if mode == "Region search":
        st.markdown("""
        <div class="sc-desc"><p>
        <strong>Region search</strong> — Search DuckDuckGo in multiple languages (EN, DE, FR, ES) for outbound travel agencies in any city.
        Each segment runs several queries to surface agencies that actively send clients to European destinations.
        Results are enriched by visiting each company website to extract email, phone, and a one-sentence description.
        </p></div>""", unsafe_allow_html=True)

        with st.form("region_form"):
            c1, c2 = st.columns([5, 3])
            with c1:
                location = st.text_input(t["location"], placeholder=t["location_placeholder"])
            with c2:
                default_segs = ["Luxury Tour Operator", "MICE Agency", "Incentive Agency", "High-End Travel Advisor"]
                segments = st.multiselect(t["segments"], list(all_segments.keys()), default=[s for s in default_segs if s in all_segments])
            custom_q = st.text_input(t["custom_query"])
            submitted = st.form_submit_button(t["run"], type="primary", use_container_width=True)

        with st.expander("Add a custom segment"):
            st.caption("Add your own segment with custom search queries. It will appear in the segments list above.")
            ns_name = st.text_input("Segment name", placeholder="e.g. Wellness Travel Agency")
            ns_queries = st.text_area("Search queries (one per line)", placeholder="wellness travel agency Europe\nWellness Reiseveranstalter Europa", height=100)
            if st.button("Add segment", type="secondary"):
                if ns_name.strip() and ns_queries.strip():
                    q_list = [q.strip() for q in ns_queries.splitlines() if q.strip()]
                    st.session_state.custom_segments[ns_name.strip()] = q_list
                    st.success(f"Added '{ns_name.strip()}' — it now appears in the segments list.")
                    st.rerun()
                else:
                    st.warning("Enter both a segment name and at least one query.")

        if submitted:
            if not location: st.error(t["err_location"]); st.stop()
            if not segments and not custom_q: st.error(t["err_segments"]); st.stop()
            queries = [(seg, q) for seg in segments for q in all_segments.get(seg, [])]
            if custom_q: queries.append(("Custom", custom_q))
            prog = st.progress(0.0); stat = st.empty()
            def cb(cur, tot, msg): prog.progress(min(cur/max(tot,1),1.0)); stat.markdown(f"**{msg}** · {cur}/{tot}")
            try:
                leads = run_pipeline(location=location, queries=queries, max_per_query=max_per_query,
                    target_keywords=target_kw, existing_partners=existing,
                    domain_blacklist=DOMAIN_BLACKLIST, enrich_websites=enrich_opt, progress_callback=cb)
                dedup_db.annotate_leads_with_history(leads); dedup_db.record_leads(leads)
                st.session_state.leads_df = pd.DataFrame(leads_to_records(leads))
                prog.empty(); stat.empty()
                st.success(f"{t['found']} {len(leads)} {t['leads_in']} {location}")
            except Exception as e:
                prog.empty(); stat.empty(); st.error(f"{type(e).__name__}: {e}")

    # --- Competitor mining ---
    elif mode == "Competitor mining":
        st.markdown("""
        <div class="sc-desc"><p>
        <strong>Competitor mining</strong> — The highest-quality lead source available.
        We scan the partner, trade, and press pages of comparable luxury hotels (Maslina, Borgo Egnazia, Aman Sveti Stefan, etc.)
        and extract the agencies they work with. Every agency listed on a competitor's site is already a <em>proven buyer</em>
        of luxury Mediterranean travel — they just don't know San Canzian yet.
        These leads convert at significantly higher rates than generic search results.
        </p></div>""", unsafe_allow_html=True)

        with st.form("comp_form"):
            urls_text = st.text_area("Hotel URLs (one per line)", value="\n".join(SUGGESTED_COMPETITORS), height=160)
            submitted = st.form_submit_button("Mine partner pages", type="primary", use_container_width=True)

        if submitted:
            urls = [u.strip() for u in urls_text.splitlines() if u.strip()]
            if not urls: st.error("Add at least one hotel URL."); st.stop()
            prog = st.progress(0.0); stat = st.empty()
            all_p = []
            for i, url in enumerate(urls):
                stat.markdown(f"**Scanning {url}**")
                try: all_p.extend(competitor_mining.mine_competitor(url))
                except Exception as e: st.warning(f"Skipped {url}: {e}")
                prog.progress((i+1)/len(urls))
            seen = set(); unique = [p for p in all_p if p["url"] not in seen and not seen.add(p["url"])]
            stat.markdown(f"**Found {len(unique)} candidates. Enriching...**")
            try:
                leads = enrich_urls(unique, target_keywords=target_kw, existing_partners=existing, source="competitor")
                dedup_db.annotate_leads_with_history(leads); dedup_db.record_leads(leads)
                st.session_state.leads_df = pd.DataFrame(leads_to_records(leads))
                prog.empty(); stat.empty()
                st.success(f"{t['found']} {len(leads)} partner candidates")
            except Exception as e:
                prog.empty(); stat.empty(); st.error(f"{type(e).__name__}: {e}")

    # --- Directory ---
    elif mode == "Directory":
        st.markdown("""
        <div class="sc-desc"><p>
        <strong>Directory scrape</strong> — Paste the URL of any public industry member listing: luxury travel network directories,
        trade association member pages, award shortlists, or "as featured in" press pages.
        The tool extracts all linked agency websites and runs them through the same enrichment pipeline.
        Note: directories behind a login wall (Virtuoso full list, ASTA) will not work — public pages only.
        </p></div>""", unsafe_allow_html=True)

        with st.form("dir_form"):
            preset = st.selectbox("Preset", ["— custom URL —"] + list(directories.KNOWN_DIRECTORIES.keys()))
            preset_url = directories.KNOWN_DIRECTORIES.get(preset, "") if preset != "— custom URL —" else ""
            dir_url = st.text_input("Directory URL", value=preset_url, placeholder="https://example.com/members")
            deep = st.checkbox("Follow pagination (up to 5 extra pages)", value=False)
            submitted = st.form_submit_button("Scrape directory", type="primary", use_container_width=True)

        if submitted:
            if not dir_url: st.error("Paste a URL."); st.stop()
            prog = st.progress(0.0); stat = st.empty()
            try:
                found = directories.scrape_directory(dir_url, deep=deep,
                    progress_callback=lambda m: stat.markdown(f"**{m}**"))
                if not found:
                    prog.empty(); stat.empty()
                    st.warning("No links found. The page may require login or use JavaScript to render content.")
                else:
                    stat.markdown(f"**{len(found)} candidates found. Enriching...**")
                    leads = enrich_urls(found, target_keywords=target_kw, existing_partners=existing, source="directory")
                    dedup_db.annotate_leads_with_history(leads); dedup_db.record_leads(leads)
                    st.session_state.leads_df = pd.DataFrame(leads_to_records(leads))
                    prog.empty(); stat.empty()
                    st.success(f"{t['found']} {len(leads)} candidates")
            except Exception as e:
                prog.empty(); stat.empty(); st.error(f"{type(e).__name__}: {e}")

    # ---------- Results ----------
    df = st.session_state.leads_df
    if df is not None and len(df) > 0:
        st.divider()
        dv = df[~df["existing_partner"]].copy()

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total leads", len(dv))
        c2.metric("Qualified", int((dv["status"]=="Qualified").sum()))
        c3.metric("Needs review", int((dv["status"]=="Needs Review").sum()))
        c4.metric("With email", int((dv["email"]!="").sum()))
        c5.metric("New", int(dv["is_new"].sum()) if "is_new" in dv else "—")
        c6.metric("FAM flagged", int(dv["attends_fam"].sum()) if "attends_fam" in dv else "—")
        st.write("")

        st.markdown("#### Results")
        f1,f2,f3,f4,f5 = st.columns(5)
        sf = f1.multiselect("Status", ["Qualified","Needs Review","Disqualified"], default=["Qualified","Needs Review"])
        segs = sorted(df["segment"].unique().tolist())
        segf = f2.multiselect("Segment", segs, default=segs)
        emailf = f3.checkbox("Email only")
        newf = f4.checkbox("New only")
        famf = f5.checkbox("FAM only")
        showex = st.checkbox(t["show_existing_partners"])

        filt = df.copy()
        if not showex: filt = filt[~filt["existing_partner"]]
        filt = filt[filt["status"].isin(sf) & filt["segment"].isin(segf)]
        if emailf: filt = filt[filt["email"]!=""]
        if newf and "is_new" in filt.columns: filt = filt[filt["is_new"]]
        if famf and "attends_fam" in filt.columns: filt = filt[filt["attends_fam"]]

        if len(filt) == 0:
            st.info(t["no_results"])
        else:
            disp = [c for c in ["status","fit_score","is_new","attends_fam","company","segment","city","country","email","phone","website","specialization"] if c in filt.columns]
            st.dataframe(filt[disp], use_container_width=True, hide_index=True, height=480,
                column_config={
                    "status": st.column_config.TextColumn("Status", width=110),
                    "fit_score": st.column_config.NumberColumn("Fit", format="%d", width=60),
                    "is_new": st.column_config.CheckboxColumn("New", width=55),
                    "attends_fam": st.column_config.CheckboxColumn("FAM", width=55),
                    "company": st.column_config.TextColumn("Company", width=200),
                    "email": st.column_config.TextColumn("Email", width=200),
                    "website": st.column_config.LinkColumn("Website"),
                    "specialization": st.column_config.TextColumn("Specialization", width=380),
                })

            def build_excel(df):
                wb = Workbook(); ws = wb.active; ws.title = "Leads"
                cols = [c for c in ["status","fit_score","is_new","attends_fam","company","segment","city","country","address","phone","email","website","specialization","latitude","longitude","source","existing_partner","previously_seen","notes"] if c in df.columns]
                pretty = {"status":"Status","fit_score":"Fit","is_new":"New?","attends_fam":"FAM?","company":"Company","segment":"Segment","city":"City","country":"Country","address":"Address","phone":"Phone","email":"Email","website":"Website","specialization":"Specialization","latitude":"Lat","longitude":"Lng","source":"Source","existing_partner":"Existing Partner?","previously_seen":"Prev. Seen","notes":"Notes"}
                ws.append([pretty.get(c,c) for c in cols])
                hf = PatternFill("solid", start_color="1B3A2D")
                for c in range(1, len(cols)+1):
                    cell = ws.cell(row=1, column=c)
                    cell.fill = hf; cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                ws.row_dimensions[1].height = 24
                qf=PatternFill("solid",start_color="D1FAE5"); rv=PatternFill("solid",start_color="FEF3C7"); dq=PatternFill("solid",start_color="FEE2E2")
                thin=Side(style="thin",color="E5E7EB"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
                for _,row in df.iterrows():
                    ws.append([row.get(c,"") for c in cols]); r=ws.max_row
                    fill={"Qualified":qf,"Needs Review":rv,"Disqualified":dq}.get(row.get("status",""))
                    for c in range(1,len(cols)+1):
                        cell=ws.cell(row=r,column=c); cell.font=Font(name="Arial",size=10)
                        cell.alignment=Alignment(vertical="top",wrap_text=True); cell.border=border
                        if c==1 and fill: cell.fill=fill
                widths=[12,7,7,7,28,22,14,12,38,18,28,32,60,11,11,12,14,18,30]
                for i,w in enumerate(widths[:len(cols)],start=1): ws.column_dimensions[get_column_letter(i)].width=w
                ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{ws.max_row}"
                buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

            st.download_button(t["download_excel"], data=build_excel(filt),
                file_name=f"sancanzian_leads_{datetime.now():%Y%m%d_%H%M}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True)

            if "latitude" in filt.columns:
                map_df = filt.dropna(subset=["latitude","longitude"])
                if len(map_df) > 0:
                    with st.expander(f"Map view — {len(map_df)} pinned"):
                        st.map(map_df.rename(columns={"latitude":"lat","longitude":"lon"}), latitude="lat", longitude="lon", size=20)

# =====================================================================
# TAB 2 — OUTREACH
# =====================================================================
with tab_outreach:
    st.markdown("""
    <div class="sc-desc"><p>
    <strong>Email templates</strong> — pre-written first-touch emails per segment and language.
    Select a template, edit it, save it. Preview it rendered against a real lead from your last search
    so you can see exactly what the sales team will send. Templates are stored as plain Markdown files
    in <code>data/templates/</code> — easy to edit directly.
    </p></div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        tpl_lang = st.selectbox("Language", ["en","de","hr"],
            format_func=lambda c: {"en":"English","de":"Deutsch","hr":"Hrvatski"}[c])
    with col2:
        available = templates.list_templates(tpl_lang)
        tpl_seg = st.selectbox("Segment", available) if available else None
        if not available:
            st.warning(f"No templates for '{tpl_lang}'. Add .md files to data/templates/{tpl_lang}/")

    if tpl_seg:
        tpl = templates.load_template(tpl_lang, tpl_seg)
        new_subj = st.text_input("Subject", value=tpl["subject"])
        new_body = st.text_area("Body", value=tpl["body"], height=360)
        if st.button("Save changes", type="secondary"):
            try: templates.save_template(tpl_lang, tpl_seg, new_subj, new_body); st.success("Saved.")
            except Exception as e: st.error(f"Save failed: {e}")
        st.divider()
        st.markdown("**Preview with a lead from your last search:**")
        df = st.session_state.leads_df
        if df is None or len(df) == 0:
            st.info("Run a search first, then preview templates against real leads here.")
        else:
            qdf = df[(df["status"]=="Qualified") & (~df["existing_partner"])]
            if len(qdf) == 0:
                st.info("No qualified leads in last search.")
            else:
                pick = st.selectbox("Lead", qdf.index.tolist(),
                    format_func=lambda i: f"{qdf.loc[i,'company']} — {qdf.loc[i,'city']}")
                rendered = templates.render({"subject":new_subj,"body":new_body}, qdf.loc[pick].to_dict(),
                    sender={"name":sender_name,"role":sender_role,"hotel_name":"San Canzian Hotel & Residences"})
                st.text_input("Rendered subject", value=rendered["subject"])
                st.text_area("Rendered body", value=rendered["body"], height=360)
                st.caption("Copy from above into your email client.")

# =====================================================================
# TAB 3 — HISTORY
# =====================================================================
with tab_history:
    st.markdown("""
    <div class="sc-desc"><p>
    <strong>History</strong> — every lead ever surfaced by the tool, stored locally in a SQLite database.
    Leads surfaced multiple times are flagged so you know which companies keep appearing in your target markets.
    Mark leads as contacted to track outreach progress across the team.
    </p></div>""", unsafe_allow_html=True)

    s = dedup_db.stats()
    h1,h2,h3 = st.columns(3)
    h1.metric("Total leads found", s["total"])
    h2.metric("Qualified", s["qualified"])
    h3.metric("Contacted", s["contacted"])
    st.write("")

    history = dedup_db.all_leads()
    if not history:
        st.info("No history yet. Run your first search in the Find Leads tab.")
    else:
        hdf = pd.DataFrame(history)
        hf1,hf2 = st.columns(2)
        hide_c = hf1.checkbox("Hide already contacted")
        only_q = hf2.checkbox("Qualified only")
        view = hdf.copy()
        if hide_c: view = view[view["contacted"]==0]
        if only_q: view = view[view["status"]=="Qualified"]
        st.dataframe(view, use_container_width=True, hide_index=True, height=460,
            column_config={
                "domain": st.column_config.TextColumn("Domain", width=180),
                "company": st.column_config.TextColumn("Company", width=200),
                "first_seen": st.column_config.TextColumn("First seen", width=140),
                "last_seen": st.column_config.TextColumn("Last seen", width=140),
                "search_count": st.column_config.NumberColumn("Times surfaced", width=100),
                "contacted": st.column_config.CheckboxColumn("Contacted", width=80),
            })
        st.divider()
        st.markdown("**Mark as contacted:**")
        mc1,mc2,mc3 = st.columns([2,3,1])
        mark_d = mc1.text_input("Domain", placeholder="genuss-touren.com")
        mark_n = mc2.text_input("Note", placeholder="emailed Maria — 05/05")
        if mc3.button("Mark"):
            if mark_d: dedup_db.mark_contacted(mark_d.strip(), mark_n.strip()); st.success("Marked."); st.rerun()

st.divider()
st.caption(t["footer"])
